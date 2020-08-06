# -*-coding:utf-8-*-
# 练习一:使用Excel写入一组数据：身高、体重
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws['A1'] = "身高"
ws['B1'] = "体重"
# 身高
height = [180,170,160,150]
# 体重
weight = [60,50,40,35]
# for循环
for i in range(len(height)):
    print(i)
    ws.cell(row=i+2, column=1, value=height[i])
    ws.cell(row=i+2, column=2, value=weight[i])
# Save the file
wb.save("sample.xlsx")