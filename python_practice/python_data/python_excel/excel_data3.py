# -*-coding:utf-8-*-
"""
练习二:
1.读取写入的Excel
2.计算是否为健康体重，如果是健康体重，则在旁边备注健康，并打印出来
健康体重计算公式：（身高cm-70）*60%
"""
# 导入模块
from openpyxl import Workbook, load_workbook
# 定义类
class PracticeExcel:
    # 定义写入的方法
    def create(self):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "身高"
        ws['B1'] = "体重"
        # 身高
        self.height = [180,170,160,150]
        # 体重
        weight = [66,50,40,35]
        # for循环
        for i in range(len(self.height)):
            ws.cell(row=i+2, column=1, value=self.height[i])
            ws.cell(row=i+2, column=2, value=weight[i])
        # Save the file
        wb.save("sample.xlsx")
    # 定义计算健康体重的方法
    def health_weight(self):
        # 读数据
        ld = load_workbook(filename="sample.xlsx")
        # 通过sheet找到对应的页签
        # 方法一
        # sheet = ld["Sheet"]
        # 方法二
        sheet = ld.active
        # 添加一列备注，查看是否为健康体重
        sheet["C1"] = "备注"
        for i in range(len(self.height)):
            # 获取身高体重的值
            height = sheet.cell(row=i+2, column=1).value
            weight = sheet.cell(row=i+2, column=2).value
            # 获取身高对应的健康体重
            health_w = (height-70)*0.6
            if weight == health_w:
                print("这是健康的体重",weight)
                sheet.cell(row=i + 2, column=3).value = "健康体重"
        # 注意：一定要保存
        ld.save("sample1.xlsx")
pe = PracticeExcel()
pe.create()
pe.health_weight()